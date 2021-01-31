import logging
import os

log = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    log.critical(
        "output.formatter_excel: openpyxl not installed, install: 'python -m pip install openpyxl'. Exiting"
    )
    raise SystemExit()

_name_map_ = {"excel_formatter": "excel"}


def excel_formatter(data, **kwargs):
    """Method to format data as an .xlsx table using openpyxl module."""
    # get arguments
    try:
        table = kwargs["table"]
    except KeyError:
        log.critical(
            "output.formatter_excel: output tag missing table definition. Exiting"
        )
        raise SystemExit()
    update = kwargs.get("update")
    url = kwargs.get("url", "./Output/")
    
    # from filename
    filename = kwargs.get("filename")
    if not filename.endswith(".xlsx"):
        filename = "{}.xlsx".format(filename)
        
    # form table_tabs - list of dictionaries
    table_tabs = []
    for index, tab_det in enumerate(table):
        tab_name = (
            tab_det.pop("tab_name")
            if "tab_name" in tab_det
            else "Sheet{}".format(index)
        )
        headers = tab_det.get("headers", None)
        if isinstance(headers, str):
            headers = [i.strip() for i in headers.split(",")]
        # get attributes out of tab_det
        tab_kwargs = {
            "path": [i.strip() for i in tab_det.get("path", "").split(".")],
            "headers": headers,
            "missing": tab_det.get("missing", ""),
            "key": tab_det.get("key", ""),
        }
        # form tab table
        tab_table_data = _ttp_["formatters"]["table"](data, **tab_kwargs)
        table_tabs.append({"name": tab_name, "data": tab_table_data})
        
    # check if need to load existing workbook
    if update and os.path.exists(os.path.join(url, filename)):
        wb = load_workbook(os.path.join(url, filename))
    # create workbook
    else:
        wb = Workbook(write_only=True)

    # add data to workbook
    for tab in table_tabs:
        # check if need to add to existing tab
        if tab["name"] in wb and update:
            ws = wb[tab["name"]]
            for row in tab["data"][1:]:
                ws.append(row)
        # create new tab
        else:
            ws = wb.create_sheet(title=tab["name"])
            for row in tab["data"]:
                ws.append(row)
    return wb
