import sys

sys.path.insert(0, "../..")
import pprint
import pytest
from shutil import copyfile

import logging

logging.basicConfig(level="ERROR")

from ttp import ttp


def test_csv_formatter_simple():
    template_1 = """
<input load="text">
interface Port-Chanel11
  vlan 10
interface Loopback0
  vlan 20
</input>

<group>
interface {{ interface }}
  vlan {{ vlan | to_int }}
</group>

<output
format="csv"
/>
"""
    parser = ttp(template=template_1)
    parser.parse()
    res = parser.result()
    assert res == ['"interface","vlan"\n"Port-Chanel11","10"\n"Loopback0","20"']


def test_csv_formatter_simple_empty_tag():
    template_1 = """
<input load="text">
interface Port-Chanel11
  vlan 10
interface Loopback0
  vlan 20
</input>

<group>
interface {{ interface }}
  vlan {{ vlan | to_int }}
</group>

<output
format="csv"
>
</output>
"""
    parser = ttp(template=template_1)
    parser.parse()
    res = parser.result()
    assert res == ['"interface","vlan"\n"Port-Chanel11","10"\n"Loopback0","20"']


# test_csv_formatter_simple_empty_tag()


def test_csv_formatter_with_is_equal():
    template = """
<input load="text" groups="interfaces2.trunks2">
interface GigabitEthernet3/3
 switchport trunk allowed vlan add 138,166-173 
 description some description
!
interface GigabitEthernet3/4
 switchport trunk allowed vlan add 100-105
!
interface GigabitEthernet3/5
 switchport trunk allowed vlan add 459,531,704-707
 ip address 1.1.1.1 255.255.255.255
 vrf forwarding ABC_VRF
!
</input>

<!--group with group specific outputs:-->
<group name="interfaces2.trunks2" output="out_csv2, test3-1">
interface {{ interface }}
 switchport trunk allowed vlan add {{ trunk_vlans }}
 description {{ description | ORPHRASE }}
 vrf forwarding {{ vrf }}
 ip address {{ ip }} {{ mask }}
!{{ _end_ }}
</group>

<out>
name="out_csv2"
path="interfaces2.trunks2"
format="csv"
sep=","
missing="undefined"
description="group specific csv outputter"
</out>

<out 
name="test3-1"
load="text"
returner="self"
functions="is_equal"
description="test csv group specific outputter"
>"description","interface","ip","mask","trunk_vlans","vrf"
"some description","GigabitEthernet3/3","undefined","undefined","138,166-173","undefined"
"undefined","GigabitEthernet3/4","undefined","undefined","100-105","undefined"
"undefined","GigabitEthernet3/5","1.1.1.1","255.255.255.255","459,531,704-707","ABC_VRF"</out>
"""
    parser = ttp(template=template)
    parser.parse()
    res = parser.result()
    assert res == [
        [
            [
                {
                    "is_equal": True,
                    "output_description": "test csv group specific outputter",
                    "output_name": "test3-1",
                }
            ]
        ]
    ]


def test_excel_formatter():
    template = """
<input load="text">
interface Loopback0
 description Router-id-loopback
 ip address 192.168.0.113/24
!
interface Loopback1
 description Router-id-loopback
 ip address 192.168.0.1/24
!
interface Vlan778
 ip address 2002::fd37/124
 ip vrf CPE1
!
interface Vlan779
 ip address 2002::bbcd/124
 ip vrf CPE2
!
</input>

<group name="loopbacks**.{{ interface }}">
interface {{ interface | contains("Loop") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<group name="vlans*">
interface {{ interface | contains("Vlan") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<output 
format="excel" 
returner="file"
filename="excel_out_test_excel_formatter"
url="./Output/"
load="yaml"
>
table:
  - headers: interface, ip, mask, vrf, description
    path: loopbacks
    key: interface
    tab_name: loopbacks
  - path: vlans
</output>
    """
    parser = ttp(template=template)
    parser.parse()
    # res = parser.result()
    # pprint.pprint(res)
    # load created workbook and test it
    from openpyxl import load_workbook

    wb = load_workbook("./Output/excel_out_test_excel_formatter.xlsx", data_only=True)
    table = {}
    for sheet_name in wb.sheetnames:
        table[sheet_name] = []
        sheet_obj = wb[sheet_name]
        for row in sheet_obj.rows:
            table[sheet_name].append([i.value for i in row])
    # pprint.pprint(table)
    assert table == {
        "Sheet1": [
            ["interface", "ip", "mask", "vrf"],
            ["Vlan778", "2002::fd37", "124", "CPE1"],
            ["Vlan779", "2002::bbcd", "124", "CPE2"],
        ],
        "loopbacks": [
            ["interface", "ip", "mask", "vrf", "description"],
            ["Loopback0", "192.168.0.113", "24", None, "Router-id-loopback"],
            ["Loopback1", "192.168.0.1", "24", None, "Router-id-loopback"],
        ],
    }


# test_excel_formatter()


def test_excel_formatter_file_returner_incomplete_url():
    template = """
<input load="text">
interface Loopback0
 description Router-id-loopback
 ip address 192.168.0.113/24
!
interface Loopback1
 description Router-id-loopback
 ip address 192.168.0.1/24
!
interface Vlan778
 ip address 2002::fd37/124
 ip vrf CPE1
!
interface Vlan779
 ip address 2002::bbcd/124
 ip vrf CPE2
!
</input>

<group name="loopbacks**.{{ interface }}">
interface {{ interface | contains("Loop") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<group name="vlans*">
interface {{ interface | contains("Vlan") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<output 
format="excel" 
returner="file"
filename="test_excel_formatter_file_returner_incomplete_url"
url="./Output"
load="yaml"
>
table:
  - headers: interface, ip, mask, vrf, description
    path: loopbacks
    key: interface
    tab_name: loopbacks
  - path: vlans
</output>
    """
    parser = ttp(template=template)
    parser.parse()
    # res = parser.result()
    # pprint.pprint(res)
    # load created workbook and test it
    from openpyxl import load_workbook

    wb = load_workbook(
        "./Output/test_excel_formatter_file_returner_incomplete_url.xlsx",
        data_only=True,
    )
    table = {}
    for sheet_name in wb.sheetnames:
        table[sheet_name] = []
        sheet_obj = wb[sheet_name]
        for row in sheet_obj.rows:
            table[sheet_name].append([i.value for i in row])
    # pprint.pprint(table)
    assert table == {
        "Sheet1": [
            ["interface", "ip", "mask", "vrf"],
            ["Vlan778", "2002::fd37", "124", "CPE1"],
            ["Vlan779", "2002::bbcd", "124", "CPE2"],
        ],
        "loopbacks": [
            ["interface", "ip", "mask", "vrf", "description"],
            ["Loopback0", "192.168.0.113", "24", None, "Router-id-loopback"],
            ["Loopback1", "192.168.0.1", "24", None, "Router-id-loopback"],
        ],
    }


# test_excel_formatter_file_returner_incomplete_url()


def test_excel_formatter_update():
    template = """
<input load="text">
interface Loopback0
 description Router-id-loopback
 ip address 192.168.0.113/24
!
interface Loopback1
 description Router-id-loopback
 ip address 192.168.0.1/24
!
interface Vlan778
 ip address 2002::fd37/124
 ip vrf CPE1
!
interface Vlan779
 ip address 2002::bbcd/124
 ip vrf CPE2
!
</input>

<group name="loopbacks_new**.{{ interface }}">
interface {{ interface | contains("Loop") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<group name="vlans*">
interface {{ interface | contains("Vlan") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<output load="yaml">
format: excel
filename: excel_out_test_excel_formatter_update.xlsx
returner: file
update: True
url: "./Output/"
table:
  - headers: interface, ip, mask, vrf, description
    path: loopbacks_new
    key: interface
    tab_name: loopbacks_new
  - path: vlans
</output>
    """
    # copy workbook to update it
    copyfile(
        src="./assets/excel_out_test_excel_formatter_update_source.xlsx",
        dst="./Output/excel_out_test_excel_formatter_update.xlsx",
    )
    # run parsing
    parser = ttp(template=template)
    parser.parse()
    # res = parser.result()
    # pprint.pprint(res)
    # load updated workbook and test it
    from openpyxl import load_workbook

    wb = load_workbook(
        "./Output/excel_out_test_excel_formatter_update.xlsx", data_only=True
    )
    table = {}
    for sheet_name in wb.sheetnames:
        table[sheet_name] = []
        sheet_obj = wb[sheet_name]
        for row in sheet_obj.rows:
            table[sheet_name].append([i.value for i in row])
    # pprint.pprint(table, width=150)
    assert table == {
        "Sheet1": [
            ["interface", "ip", "mask", "vrf"],
            ["Vlan778", "2002::fd37", "124", "CPE1"],
            ["Vlan779", "2002::bbcd", "124", "CPE2"],
            ["Vlan778", "2002::fd37", "124", "CPE1"],  # << appended to existing tab
            ["Vlan779", "2002::bbcd", "124", "CPE2"],
        ],  # << appended to existing tab
        "loopbacks": [
            [
                "interface",
                "ip",
                "mask",
                "vrf",
                "description",
            ],  # << existing tab unchanged
            ["Loopback0", "192.168.0.113", "24", None, "Router-id-loopback"],
            ["Loopback1", "192.168.0.1", "24", None, "Router-id-loopback"],
        ],
        "loopbacks_new": [
            ["interface", "ip", "mask", "vrf", "description"],  # << new tab created
            ["Loopback0", "192.168.0.113", "24", None, "Router-id-loopback"],
            ["Loopback1", "192.168.0.1", "24", None, "Router-id-loopback"],
        ],
    }


# test_excel_formatter_update()


def test_excel_formatter_update_using_result_kwargs():
    template = """
<input load="text">
interface Loopback0
 description Router-id-loopback
 ip address 192.168.0.113/24
!
interface Loopback1
 description Router-id-loopback
 ip address 192.168.0.1/24
!
interface Vlan778
 ip address 2002::fd37/124
 ip vrf CPE1
!
interface Vlan779
 ip address 2002::bbcd/124
 ip vrf CPE2
!
</input>

<group name="loopbacks_new**.{{ interface }}">
interface {{ interface | contains("Loop") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<group name="vlans*">
interface {{ interface | contains("Vlan") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>
    """
    # copy workbook to update it
    copyfile(
        src="./assets/excel_out_test_excel_formatter_update_source.xlsx",
        dst="./Output/excel_out_test_excel_formatter_update.xlsx",
    )
    # run parsing
    parser = ttp(template=template)
    parser.parse()
    # form excel table and save in file
    parser.result(
        format="excel",
        filename="excel_out_test_excel_formatter_update.xlsx",
        returner="file",
        update=True,
        url="./Output/",
        table=[
            {
                "headers": ["interface", "ip", "mask", "vrf", "description"],
                "path": "loopbacks_new",
                "key": "interface",
                "tab_name": "loopbacks_new",
            },
            {"path": "vlans"},
        ],
    )
    # pprint.pprint(res)
    # load updated workbook and test it
    from openpyxl import load_workbook

    wb = load_workbook(
        "./Output/excel_out_test_excel_formatter_update.xlsx", data_only=True
    )
    table = {}
    for sheet_name in wb.sheetnames:
        table[sheet_name] = []
        sheet_obj = wb[sheet_name]
        for row in sheet_obj.rows:
            table[sheet_name].append([i.value for i in row])
    # pprint.pprint(table, width=150)
    assert table == {
        "Sheet1": [
            ["interface", "ip", "mask", "vrf"],
            ["Vlan778", "2002::fd37", "124", "CPE1"],
            ["Vlan779", "2002::bbcd", "124", "CPE2"],
            ["Vlan778", "2002::fd37", "124", "CPE1"],  # << appended to existing tab
            ["Vlan779", "2002::bbcd", "124", "CPE2"],
        ],  # << appended to existing tab
        "loopbacks": [
            [
                "interface",
                "ip",
                "mask",
                "vrf",
                "description",
            ],  # << existing tab unchanged
            ["Loopback0", "192.168.0.113", "24", None, "Router-id-loopback"],
            ["Loopback1", "192.168.0.1", "24", None, "Router-id-loopback"],
        ],
        "loopbacks_new": [
            ["interface", "ip", "mask", "vrf", "description"],  # << new tab created
            ["Loopback0", "192.168.0.113", "24", None, "Router-id-loopback"],
            ["Loopback1", "192.168.0.1", "24", None, "Router-id-loopback"],
        ],
    }


# test_excel_formatter_update_using_result_kwargs()


def test_excel_formatter_no_results_at_path_strict_false():
    template = """
<input load="text">
interface Loopback0
 description Router-id-loopback
 ip address 192.168.0.113/24
!
interface Loopback1
 description Router-id-loopback
 ip address 192.168.0.1/24
</input>

<group name="loopbacks**.{{ interface }}">
interface {{ interface | contains("Loop") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<output 
format="excel" 
returner="file"
filename="test_excel_formatter_no_results_at_path_strict_false"
url="./Output/"
load="yaml"
>
table:
  - headers: interface, ip, mask, vrf, description
    path: loopbacks
    key: interface
    tab_name: loopbacks
  - path: vlans
    strict: False
</output>
    """
    parser = ttp(template=template)
    parser.parse()
    # res = parser.result()
    # pprint.pprint(res)
    # load created workbook and test it
    from openpyxl import load_workbook

    wb = load_workbook(
        "./Output/test_excel_formatter_no_results_at_path_strict_false.xlsx",
        data_only=True,
    )
    table = {}
    for sheet_name in wb.sheetnames:
        table[sheet_name] = []
        sheet_obj = wb[sheet_name]
        for row in sheet_obj.rows:
            table[sheet_name].append([i.value for i in row])
    # pprint.pprint(table)
    assert table == {
        "Sheet1": [],
        "loopbacks": [
            ["interface", "ip", "mask", "vrf", "description"],
            ["Loopback0", "192.168.0.113", "24", None, "Router-id-loopback"],
            ["Loopback1", "192.168.0.1", "24", None, "Router-id-loopback"],
        ],
    }


# test_excel_formatter_no_results_at_path_strict_false()


def test_excel_formatter_no_results_at_path_strict_true():
    """
    This test should raise key-error exception within traverse function
    as path: vlans not found in results
    """
    template = """
<input load="text">
interface Loopback0
 description Router-id-loopback
 ip address 192.168.0.113/24
!
interface Loopback1
 description Router-id-loopback
 ip address 192.168.0.1/24
</input>

<group name="loopbacks**.{{ interface }}">
interface {{ interface | contains("Loop") }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<output 
format="excel" 
returner="file"
filename="test_excel_formatter_no_results_at_path_strict_true"
url="./Output/"
load="yaml"
>
table:
  - headers: interface, ip, mask, vrf, description
    path: loopbacks
    key: interface
    tab_name: loopbacks
  - path: vlans
    strict: True
</output>
    """
    parser = ttp(template=template)
    with pytest.raises(KeyError):
        parser.parse()


# test_excel_formatter_no_results_at_path_strict_true()


def test_tabulate_formatter():
    template = """
<input load="text">
router bgp 65100
  neighbor 10.145.1.9
    description vic-mel-core1
  !
  neighbor 192.168.101.1
    description qld-bri-core1
</input>

<group name="bgp_config">
router bgp {{ bgp_as }}
 <group name="peers">
  neighbor {{ peer }}
    description {{ description  }}
 </group>
</group> 
    
<output 
path="bgp_config.peers" 
format="tabulate" 
format_attributes="tablefmt='fancy_grid'"
/>
    """
    parser = ttp(template=template)
    parser.parse()
    res = parser.result()
    # pprint.pprint(res)
    assert res == [
        "╒═══════════════╤═══════════════╕\n"
        "│ description   │ peer          │\n"
        "╞═══════════════╪═══════════════╡\n"
        "│ vic-mel-core1 │ 10.145.1.9    │\n"
        "├───────────────┼───────────────┤\n"
        "│ qld-bri-core1 │ 192.168.101.1 │\n"
        "╘═══════════════╧═══════════════╛"
    ]


# test_tabulate_formatter()


def test_table_formatter():
    template = """
<input load="text">
router bgp 65100
  neighbor 10.145.1.9
    description vic-mel-core1
  !
  neighbor 192.168.101.1
    description qld-bri-core1
</input>

<group name="bgp_config">
router bgp {{ bgp_as }}
 <group name="peers">
  neighbor {{ peer }}
    description {{ description  }}
 </group>
</group> 
    
<output 
path="bgp_config.peers" 
format="table" 
headers="description, peer, asn"
missing="Undefined"
/>
    """
    parser = ttp(template=template)
    parser.parse()
    res = parser.result()
    # pprint.pprint(res)
    assert res == [
        [
            ["description", "peer", "asn"],
            ["vic-mel-core1", "10.145.1.9", "Undefined"],
            ["qld-bri-core1", "192.168.101.1", "Undefined"],
        ]
    ]


# test_table_formatter()


def test_table_formatter_with_key():
    template = """
<input load="text">
interface Loopback0
 description Router-id-loopback
 ip address 192.168.0.113/24
!
interface Loopback1
 description Router-id-loopback
 ip address 192.168.0.1/24
!
interface Vlan778
 ip address 2002::fd37/124
 ip vrf CPE1
!
interface Vlan779
 ip address 2002::bbcd/124
 ip vrf CPE2
!
</input>
<group name="interfaces**.{{ interface }}">
interface {{ interface }}
 ip address {{ ip }}/{{ mask }}
 description {{ description }}
 ip vrf {{ vrf }}
</group>

<output 
path="interfaces" 
format="table" 
headers="intf, ip, mask, vrf, description, switchport"
key="intf"
missing="Undefined"
/>
    """
    parser = ttp(template=template)
    parser.parse()
    res = parser.result()
    # pprint.pprint(res, width=100)
    assert res == [
        [
            ["intf", "ip", "mask", "vrf", "description", "switchport"],
            [
                "Loopback0",
                "192.168.0.113",
                "24",
                "Undefined",
                "Router-id-loopback",
                "Undefined",
            ],
            [
                "Loopback1",
                "192.168.0.1",
                "24",
                "Undefined",
                "Router-id-loopback",
                "Undefined",
            ],
            ["Vlan778", "2002::fd37", "124", "CPE1", "Undefined", "Undefined"],
            ["Vlan779", "2002::bbcd", "124", "CPE2", "Undefined", "Undefined"],
        ]
    ]


# test_table_formatter_with_key()


def test_table_formatter_multiple_input():
    template = """
<input load="text">
interface Loopback0
 ip address 192.168.0.113/24
!
interface Vlan778
 ip address 2002::fd37/124
!
</input>

<input load="text">
interface Loopback10
 ip address 192.168.0.10/24
!
interface Vlan710
 ip address 2002::fd10/124
!
</input>

<group>
interface {{ interface }}
 ip address {{ ip }}/{{ mask }}
</group>

<output format="table"/>
    """
    parser = ttp(template=template)
    parser.parse()
    res = parser.result()
    # pprint.pprint(res, width=100)
    assert res == [
        [
            ["interface", "ip", "mask"],
            ["Loopback0", "192.168.0.113", "24"],
            ["Vlan778", "2002::fd37", "124"],
            ["Loopback10", "192.168.0.10", "24"],
            ["Vlan710", "2002::fd10", "124"],
        ]
    ]


# test_table_formatter_multiple_input()


def test_jinja2_formatter():
    template = """
<input load="text">
interface Loopback0
 ip address 192.168.0.113/24
!
interface Vlan778
 ip address 2002::fd37/124
!
</input>

<input load="text">
interface Loopback10
 ip address 192.168.0.10/24
!
interface Vlan710
 ip address 2002::fd10/124
!
</input>

<group>
interface {{ interface }}
 ip address {{ ip }}/{{ mask }}
</group>

<output format="jinja2">{{ _data_ }}</output>
    """
    parser = ttp(template=template)
    parser.parse()
    res = parser.result()
    # pprint.pprint(res, width=100)
    assert res == [
        "[[{'ip': '192.168.0.113', 'mask': '24', 'interface': 'Loopback0'}, {'ip': '2002::fd37', 'mask': "
        "'124', 'interface': 'Vlan778'}], [{'ip': '192.168.0.10', 'mask': '24', 'interface': "
        "'Loopback10'}, {'ip': '2002::fd10', 'mask': '124', 'interface': 'Vlan710'}]]"
    ]


# test_jinja2_formatter()
